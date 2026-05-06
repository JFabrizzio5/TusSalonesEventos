import os

def create_database_config(project_path, db_type):
    """Crea la carpeta config/ y el archivo database.py con lógica de descubrimiento dinámico."""
    if db_type == "none":
        return

    config_path = os.path.join(project_path, "config")
    os.makedirs(config_path, exist_ok=True)
    if db_type == "excel":
        storage_path = os.path.join(project_path, "storage")
        os.makedirs(storage_path, exist_ok=True)
        gitkeep = os.path.join(storage_path, ".gitkeep")
        if not os.path.exists(gitkeep):
            with open(gitkeep, "w") as f:
                f.write("")
    
    db_content = """import asyncio
import os
from fastapi import HTTPException

# Carga de Ambiente Dynamica (dev / prod / docker)
ENVIRONMENT = os.getenv("ENVIRONMENT", "dev")

async def get_db_url(db_type: str) -> str:
    \"\"\"Devuelve la configuración local (dev/prod) o llama al IAM si no está definida.\"\"\"
    env_suffix = f"_{ENVIRONMENT.upper()}" if ENVIRONMENT else "_DEV"
    if db_type == "sql":
        url = os.getenv(f"DATABASE_URL{env_suffix}")
        if url: return url
        return os.getenv("DATABASE_URL_DEV") # Fallback local
    if db_type == "excel":
        url = os.getenv(f"EXCEL_STORAGE_PATH{env_suffix}") or os.getenv("EXCEL_STORAGE_PATH")
        if url: return url
        return os.path.join("storage", "data.xlsx")
    else:
        url = os.getenv(f"MONGODB_URL{env_suffix}")
        if url: return url
        return os.getenv("MONGODB_URL_DEV")

async def check_db_health() -> dict:
    \"\"\"Verifica la conectividad real con las bases de datos configuradas.\"\"\"
    results = {}
"""

    if db_type in ["sql", "both"]:
        db_content += """
    # Verificar SQL
    try:
        from sqlalchemy import text
        await init_sql_engine()
        async with _AsyncSessionLocal() as session:
            await session.execute(text("SELECT 1"))
        results["sql"] = "Conectado"
    except Exception as e:
        results["sql"] = f"Error: {str(e)}"
"""

    if db_type in ["mongo", "both"]:
        db_content += """
    # Verificar Mongo
    try:
        await init_mongo_client()
        await _mongo_client.admin.command('ping')
        results["mongo"] = "Conectado"
    except Exception as e:
        results["mongo"] = f"Error: {str(e)}"
"""

    if db_type == "excel":
        db_content += """
    # Verificar Excel
    try:
        path = await get_excel_path()
        results["excel"] = f"Archivo listo: {path}"
    except Exception as e:
        results["excel"] = f"Error: {str(e)}"
"""

    db_content += """    return results

Base = None
"""

    if db_type in ["sql", "both"]:
        db_content += """
# ==========================================
# Configuración PostgreSQL (SQLAlchemy)
# ==========================================
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker, declarative_base

# Base para todos los modelos del proyecto (Necesaria para Alembic)
Base = declarative_base()

_sql_engine = None
_AsyncSessionLocal = None

async def init_sql_engine():
    global _sql_engine, _AsyncSessionLocal
    if not _sql_engine:
        url = await get_db_url("sql")
        _sql_engine = create_async_engine(
            url, 
            echo=(ENVIRONMENT == "dev"),
            pool_size=20
        )
        
        # Activar soporte vectorial en PostgreSQL automáticamente
        from sqlalchemy import text
        try:
            async with _sql_engine.begin() as conn:
                await conn.execute(text("CREATE EXTENSION IF NOT EXISTS vector"))
        except Exception as e:
            pass # Si el usuario no tiene permisos o ya existe, omitir
            
        _AsyncSessionLocal = sessionmaker(
            bind=_sql_engine, class_=AsyncSession, expire_on_commit=False
        )

async def get_sql_db(tenant_id: str = None):
    await init_sql_engine()
    async with _AsyncSessionLocal() as session:
        if tenant_id:
            from sqlalchemy import text
            import re
            # Validación estricta para evitar SQL Injection en el search_path
            if not re.match(r"^[a-zA-Z0-9_]+$", tenant_id):
                raise ValueError("Formato de tenant_id inválido")
            await session.execute(text(f'SET search_path TO "tenant_{tenant_id}"'))
        yield session
"""

    if db_type in ["mongo", "both"]:
        db_content += """
# ==========================================
# Configuración MongoDB (Motor)
# ==========================================
from motor.motor_asyncio import AsyncIOMotorClient

_mongo_client = None

async def init_mongo_client():
    global _mongo_client
    if not _mongo_client:
        url = await get_db_url("mongo")
        _mongo_client = AsyncIOMotorClient(url, maxPoolSize=50)

async def get_mongo_db(tenant_id: str = None):
    \"\"\"Obtiene la base de datos de MongoDB con soporte para multi-tenancy.\"\"\"
    await init_mongo_client()
    db_name = os.getenv("MONGODB_DB_NAME", "saas_db")
    if tenant_id:
        return _mongo_client[f"tenant_{tenant_id}"]
    return _mongo_client[db_name]
"""

    if db_type == "excel":
        db_content += """
# ==========================================
# Configuración Excel (openpyxl)
# ==========================================
from openpyxl import Workbook, load_workbook

def _is_sheet_empty(sheet):
    return sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None

def _ensure_sheet(workbook, sheet_name, headers=None):
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    if headers and _is_sheet_empty(sheet):
        sheet.append(headers)
    return sheet

async def get_excel_path():
    path = await get_db_url("excel")
    path = path or os.path.join("storage", "data.xlsx")
    directory = os.path.dirname(path) or "."
    os.makedirs(directory, exist_ok=True)

    if not os.path.exists(path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = os.getenv("EXCEL_SHEET_DEFAULT", "sheet1")
        workbook.save(path)
    return path

async def get_excel_workbook():
    path = await get_excel_path()
    workbook = await asyncio.to_thread(load_workbook, path)
    return workbook, path

async def ensure_excel_sheet(sheet_name, headers=None):
    workbook, path = await get_excel_workbook()
    _ensure_sheet(workbook, sheet_name, headers)
    await asyncio.to_thread(workbook.save, path)
    return {"sheet": sheet_name, "path": path}

async def list_excel_rows(sheet_name):
    workbook, _ = await get_excel_workbook()
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [header for header in rows[0] if header is not None]
    if not headers:
        return []

    records = []
    for values in rows[1:]:
        if values is None or all(value is None for value in values[:len(headers)]):
            continue
        records.append({
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
        })
    return records

async def append_excel_row(sheet_name, row):
    workbook, path = await get_excel_workbook()
    headers = list(row.keys())
    sheet = _ensure_sheet(workbook, sheet_name, headers)
    current_headers = [cell.value for cell in sheet[1] if cell.value is not None]

    if not current_headers:
        sheet.append(headers)
        current_headers = headers

    sheet.append([row.get(header) for header in current_headers])
    await asyncio.to_thread(workbook.save, path)
    return row
"""
    
    with open(os.path.join(config_path, "database.py"), "w") as f:
        f.write(db_content)

def create_redis_config(project_path):
    """Genera centralizadamente la configuración para Redes en config/redis_config.py."""
    config_path = os.path.join(project_path, "config")
    os.makedirs(config_path, exist_ok=True)
    
    # Maneja la conexión individual a cada base de datos lógica de Redis
    redis_content = """import os
from arq.connections import RedisSettings

# URLs especializadas para evitar ruido entre servicios
REDIS_CACHE_URL = os.getenv("REDIS_CACHE_URL", "redis://redis:6379/0")
REDIS_PUBSUB_URL = os.getenv("REDIS_PUBSUB_URL", "redis://redis:6379/1")
REDIS_QUEUE_URL = os.getenv("REDIS_QUEUE_URL", "redis://redis:6379/2")

def get_redis_settings(url_type="queue"):
    \"\"\"Parsea el DSN de Redis para devolver la configuración compatible con ARQ/Otras.\"\"\"
    from urllib.parse import urlparse
    import os
    
    # Lectura dinámica para asegurar que load_dotenv() ya haya corrido
    c_url = os.getenv("REDIS_CACHE_URL", "redis://redis:6379/0")
    p_url = os.getenv("REDIS_PUBSUB_URL", "redis://redis:6379/1")
    q_url = os.getenv("REDIS_QUEUE_URL", "redis://redis:6379/2")
    
    url_str = q_url if url_type == "queue" else p_url
    if url_type == "cache": url_str = c_url
    
    url = urlparse(url_str)
    return RedisSettings(
        host=url.hostname or 'redis',
        port=url.port or 6379,
        database=int(url.path.strip('/') or 0),
        password=url.password
    )
"""
    with open(os.path.join(config_path, "redis_config.py"), "w") as f:
        f.write(redis_content)
